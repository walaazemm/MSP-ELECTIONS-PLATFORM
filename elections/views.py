from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count, Sum, Case, When, Value, IntegerField, Max, F, Subquery, OuterRef
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.cache import cache  # 🚀 PILLAR 2: REDIS CACHE IMPORT
import json
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import math
from territories.models import Wilaya, Commune
from .models import Election, BureauDeVote, ElectionList, StationResult, ListResult, ResultAuditLog
from django.contrib.auth import get_user_model
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

def api_communes_by_wilaya(request, wilaya_id):
    """API endpoint for dynamic commune loading (Bulletproof against duplicates)"""
    communes = Commune.objects.filter(wilaya_id=wilaya_id).order_by('name_fr').distinct()
    seen_names = set()
    unique_communes = []
    for c in communes:
        if c.name_fr not in seen_names:
            seen_names.add(c.name_fr)
            unique_communes.append({'id': c.id, 'name_fr': c.name_fr, 'name_ar': c.name_ar or ''})
    return JsonResponse(unique_communes, safe=False)

@login_required
def create_bureau(request):
    # 1. Role & Scope Setup
    try:
        role_obj = request.user.role_profile
        role = role_obj.role
    except AttributeError:
        return redirect('login')

    # 🛡️ BLOCK COMMUNE USERS (They already have 1 PV assigned to them)
    if role == 'commune':
        messages.info(request, " بصفتك ممثل بلدية، لديك محضر واحد مخصص مسبقاً. يرجى الدخول إليه من لوحة التحكم لإدخال النتائج.")
        return redirect('home')

    if role not in ['national', 'super_admin', 'wilaya']:
        messages.error(request, "غير مصرح لك بالوصول")
        return redirect('home')

    if role in ['national', 'super_admin']:
        wilayas = Wilaya.objects.all().order_by('code')
    elif role == 'wilaya':
        wilayas = Wilaya.objects.filter(id=role_obj.wilaya_id)
    else:
        wilayas = Wilaya.objects.filter(id=role_obj.commune.wilaya_id)

    elections = Election.objects.all()

    if request.method == 'POST':
        election_id = request.POST.get('election')
        commune_id = request.POST.get('commune')
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        reg_voters = request.POST.get('registered_voters', '').strip()

        if role == 'commune':
            commune_id = str(role_obj.commune_id)

        if not election_id or not commune_id or not code:
            messages.error(request, "يرجى ملء الحقول المطلوبة: الانتخاب، البلدية، رمز المحضر")
            return redirect('elections:create_bureau')

        if role in ['national', 'super_admin']:
            commune = Commune.objects.filter(id=commune_id).first()
        elif role == 'wilaya':
            commune = Commune.objects.filter(id=commune_id, wilaya_id=role_obj.wilaya_id).first()
        else:
            commune = Commune.objects.filter(id=commune_id).first()

        if not commune:
            messages.error(request, "البلدية المحددة غير صحيحة أو خارج نطاق صلاحياتك")
            return redirect('elections:create_bureau')

        election = Election.objects.filter(id=election_id).first()
        if not election:
            messages.error(request, "الانتخاب المحدد غير موجود")
            return redirect('elections:create_bureau')

        if BureauDeVote.objects.filter(commune=commune, election=election, code=code, is_deleted=False).exists():
            messages.warning(request, f"يوجد محضر برمز {code} بالفعل في هذه البلدية والانتخاب")
            return redirect('elections:create_bureau')

        reg_int = None
        if reg_voters:
            try:
                reg_int = int(reg_voters)
            except ValueError:
                messages.error(request, "عدد الناخبين المسجلين يجب أن يكون رقماً صحيحاً")
                return redirect('elections:create_bureau')

        BureauDeVote.objects.create(
            commune=commune, election=election, code=code, name=name if name else None,
            registered_voters=reg_int, created_by=request.user
        )
        messages.success(request, f" تم إضافة المحضر بنجاح (الرمز: {code})")
        return redirect('elections:create_bureau')

    return render(request, 'elections/create_bureau.html', {
        'wilayas': wilayas, 'elections': elections, 'role': role, 'role_obj': role_obj, 'active_page': 'create'
    })

@login_required
def enter_results(request, bureau_id):
    try:
        bureau = BureauDeVote.objects.select_related('commune__wilaya', 'election').get(id=bureau_id, is_deleted=False)
    except BureauDeVote.DoesNotExist:
        messages.error(request, "المحضر غير موجود")
        return redirect('home')

    try:
        role_obj = request.user.role_profile
        role = role_obj.role
        if role == 'commune' and bureau.commune.id != role_obj.commune.id:
            messages.error(request, "ليس لديك صلاحية للوصول إلى هذا المحضر")
            return redirect('home')
        elif role == 'wilaya' and bureau.commune.wilaya.id != role_obj.wilaya.id:
            messages.error(request, "ليس لديك صلاحية للوصول إلى هذا المحضر")
            return redirect('home')
    except AttributeError:
        return redirect('login')

    election = bureau.election
    current_wilaya = bureau.commune.wilaya
    
    lists = list(election.lists.filter(
        Q(wilaya__isnull=True) | Q(wilaya=current_wilaya), is_active=True
    ).order_by('display_order'))

    result = StationResult.objects.filter(election=election, bureau=bureau, is_deleted=False).first()
    if result:
        list_votes = {lr.election_list_id: lr.votes for lr in ListResult.objects.filter(station_result=result)}
        for lst in lists:
            lst.current_votes = list_votes.get(lst.id, 0)
    else:
        for lst in lists:
            lst.current_votes = 0

    wilaya_total_seats = current_wilaya.total_seats if hasattr(current_wilaya, 'total_seats') and current_wilaya.total_seats else 4
    wilaya_bureaux = BureauDeVote.objects.filter(commune__wilaya=current_wilaya, election=election, is_deleted=False)
    wilaya_results = StationResult.objects.filter(bureau__in=wilaya_bureaux, is_deleted=False)
    
    wilaya_aggr = wilaya_results.aggregate(
        total_present=Sum('total_votes_cast', default=0),
        total_null=Sum('null_votes', default=0)
    )
    w_valid_votes = (wilaya_aggr['total_present'] or 0) - (wilaya_aggr['total_null'] or 0)
    
    wilaya_list_votes = ListResult.objects.filter(
        station_result__bureau__in=wilaya_bureaux,
        station_result__is_deleted=False
    ).values('election_list_id').annotate(total=Sum('votes'))
    
    w_dict = {item['election_list_id']: item['total'] for item in wilaya_list_votes}
    
    calc_result = calculate_seats_allocation(w_valid_votes, w_dict, wilaya_total_seats)
    seats_projection = calc_result['lists']
    seats_meta = calc_result['meta']
    
    all_lists_in_wilaya = ElectionList.objects.filter(id__in=w_dict.keys()).values('id', 'name_ar', 'is_our_party')
    for lst in all_lists_in_wilaya:
        lid = lst['id']
        if lid in seats_projection:
            seats_projection[lid]['name'] = lst['name_ar']
            seats_projection[lid]['is_our_party'] = lst['is_our_party']

    context_data = {
        'bureau': bureau, 'lists': lists, 'result': result, 'active_page': 'data_entry',
        'seats_projection': seats_projection, 'seats_meta': seats_meta,
        'wilaya_total_seats': wilaya_total_seats, 'current_wilaya': current_wilaya
    }

    if request.method == 'POST':
        reg_voters = int(request.POST.get('registered_voters') or 0)
        total_present = int(request.POST.get('total_votes_cast') or 0)
        null_v = int(request.POST.get('null_votes') or 0)
        valid_votes = total_present - null_v 

        list_votes = {}
        for lst in lists:
            value = request.POST.get(f'list_{lst.id}', '').strip()
            list_votes[lst.id] = int(value) if value else 0

        sum_lists = sum(list_votes.values())
        errors = []

        if total_present > reg_voters:
            errors.append("عدد الحاضرين لا يمكن أن يتجاوز عدد الناخبين المسجلين")
        if null_v > total_present:
            errors.append("الأصوات الملغاة والبيضاء لا يمكن أن تتجاوز عدد الحاضرين")
        if valid_votes < 0:
            errors.append("الأصوات المعبر عنها لا يمكن أن تكون سالبة")
        if any(v < 0 for v in list_votes.values()):
            errors.append("لا يمكن أن تكون أصوات القوائم سالبة")
        if sum_lists != valid_votes:
            errors.append(f"⛔ خطأ في المجموع: مجموع أصوات القوائم ({sum_lists}) يجب أن يساوي تماماً الأصوات المعبر عنها ({valid_votes}).")

        if errors:
            context_data['errors'] = errors
            return render(request, 'elections/data_entry.html', context_data)

        if reg_voters != bureau.registered_voters:
            BureauDeVote.objects.filter(id=bureau.id).update(registered_voters=reg_voters)
            bureau.registered_voters = reg_voters

        with transaction.atomic():
            res, created = StationResult.objects.update_or_create(
                election=election, bureau=bureau,
                defaults={
                    'registered_voters': reg_voters, 'total_votes_cast': total_present,
                    'null_votes': null_v, 'blank_votes': 0, 'last_edited_by': request.user, 'is_deleted': False
                }
            )
            if created:
                res.submitted_by = request.user
                res.save()

            for lst_id, votes in list_votes.items():
                ListResult.objects.update_or_create(
                    station_result=res, election_list_id=lst_id, defaults={'votes': votes}
                )
                
        messages.success(request, " تم حفظ النتائج بنجاح")
        
        # 🔴 WEBSOCKET: Push real-time update to all connected dashboards
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            'dashboard_live',
            {
                'type': 'result_submitted',
                'data': {
                    'bureau_code': bureau.code,
                    'commune_name': bureau.commune.name_ar,
                    'wilaya_name': bureau.commune.wilaya.name_ar,
                    'wilaya_code': str(bureau.commune.wilaya.code).zfill(2),  # 🗺️ FIX: Forces "01" instead of 1 for SVG matching
                    'submitted_by': request.user.full_name if hasattr(request.user, 'full_name') else str(request.user),
                    'timestamp': timezone.now().strftime('%H:%M:%S'),
                    'total_votes': total_present,
                    'msp_votes': list_votes.get(
                        ElectionList.objects.filter(is_our_party=True).values_list('id', flat=True).first(), 0
                    ),
                    'message': f'🗳️ {bureau.commune.name_ar} - {bureau.code} submitted'
                }
            }
        )
        
        # 🧹 PILLAR 2: INSTANTLY CLEAR CACHE SO DASHBOARD UPDATES IN REAL-TIME
        election_id = bureau.election.id
        wilaya_id = bureau.commune.wilaya.id
        
        cache.delete(f"wilaya_math_{wilaya_id}_{election_id}")
        cache.delete(f"wilaya_detail_math_{wilaya_id}_{election_id}")
        cache.delete(f"national_math_{election_id}")
        
        return redirect('elections:enter_results', bureau_id=bureau.id)

    return render(request, 'elections/data_entry.html', context_data)

@login_required
def bureau_list(request):
    try:
        role_obj = request.user.role_profile
        role = role_obj.role
    except AttributeError:
        return redirect('login')

    bureaux = BureauDeVote.objects.select_related('commune__wilaya', 'election').filter(is_deleted=False)

    if role == 'commune':
        bureaux = bureaux.filter(commune=role_obj.commune)
    elif role == 'wilaya':
        bureaux = bureaux.filter(commune__wilaya=role_obj.wilaya)

    election_id = request.GET.get('election')
    search_query = request.GET.get('search', '').strip()
    wilaya_id = request.GET.get('wilaya')
    commune_id = request.GET.get('commune')
    status_filter = request.GET.get('status')

    if election_id:
        bureaux = bureaux.filter(election_id=election_id)
        active_election = Election.objects.filter(id=election_id).first()
    else:
        active_election = Election.objects.filter(status='open').order_by('-election_date').first()
        if active_election:
            bureaux = bureaux.filter(election=active_election)

    if search_query:
        bureaux = bureaux.filter(Q(code__icontains=search_query) | Q(name__icontains=search_query))

    if wilaya_id:
        bureaux = bureaux.filter(commune__wilaya_id=wilaya_id)
    if commune_id:
        bureaux = bureaux.filter(commune_id=commune_id)

    msp_subquery = ListResult.objects.filter(
        station_result__bureau=OuterRef('id'),
        station_result__is_deleted=False,
        election_list__is_our_party=True
    ).values('station_result__bureau').annotate(
        total_msp=Sum('votes')
    ).values('total_msp')

    bureaux = bureaux.annotate(
        has_result=Case(When(result__isnull=False, then=Value(1)), default=Value(0), output_field=IntegerField()),
        msp_votes_raw=Subquery(msp_subquery, output_field=IntegerField())
    ).annotate(
        msp_votes=Case(When(msp_votes_raw__isnull=True, then=Value(0)), default=F('msp_votes_raw'), output_field=IntegerField())
    )

    if status_filter == 'completed':
        bureaux = bureaux.filter(has_result=1)
    elif status_filter == 'pending':
        bureaux = bureaux.filter(has_result=0)

    bureaux = bureaux.distinct().order_by('commune__wilaya__code', 'commune__code', 'code')

    elections = Election.objects.all().order_by('-election_date')
    
    if role in ['national', 'super_admin']:
        wilayas = Wilaya.objects.all().order_by('name_fr')
    elif role == 'wilaya':
        wilayas = Wilaya.objects.filter(id=role_obj.wilaya_id)
    else:
        wilayas = Wilaya.objects.filter(id=role_obj.commune.wilaya_id)

    total_bureaux = bureaux.count()
    reported_bureaux = bureaux.filter(has_result=1).count()
    pending_bureaux = total_bureaux - reported_bureaux
    completion_percentage = round((reported_bureaux / total_bureaux * 100), 1) if total_bureaux > 0 else 0

    paginator = Paginator(bureaux, 50) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'elections/bureau_list.html', {
        'page_obj': page_obj, 'elections': elections, 'active_election': active_election, 'active_page': 'bureaux',
        'total_bureaux': total_bureaux, 'reported_bureaux': reported_bureaux, 'pending_bureaux': pending_bureaux,
        'completion_percentage': completion_percentage, 'wilayas': wilayas, 'search_query': search_query,
        'selected_wilaya': wilaya_id, 'selected_commune': commune_id, 'selected_status': status_filter, 'role': role,
    })

@login_required
def bulk_delete_bureaux(request):
    """Handles bulk deletion of multiple bureaux via AJAX"""
    if request.method == 'POST':
        try:
            role_obj = request.user.role_profile
            role = role_obj.role
        except AttributeError:
            return JsonResponse({'success': False, 'error': 'غير مصرح'})

        bureau_ids = request.POST.getlist('bureau_ids')
        if not bureau_ids:
            return JsonResponse({'success': False, 'error': 'لم يتم اختيار أي مركز'})

        bureaux_to_delete = BureauDeVote.objects.filter(id__in=bureau_ids)
        
        # Scope Check (Ensure users only delete within their jurisdiction)
        if role == 'commune':
            bureaux_to_delete = bureaux_to_delete.filter(commune=role_obj.commune)
        elif role == 'wilaya':
            bureaux_to_delete = bureaux_to_delete.filter(commune__wilaya=role_obj.wilaya)
        
        deleted_count = 0
        for bureau in bureaux_to_delete:
            # 🧹 HARD DELETE CASCADE
            StationResult.objects.filter(bureau=bureau).delete()
            bureau.delete()
            deleted_count += 1
            
        return JsonResponse({'success': True, 'deleted': deleted_count})
        
    return JsonResponse({'success': False, 'error': 'طلب غير صالح'})

@login_required
def edit_bureau(request, bureau_id):
    try:
        bureau = BureauDeVote.objects.select_related('commune__wilaya', 'election').get(id=bureau_id, is_deleted=False)
    except BureauDeVote.DoesNotExist:
        messages.error(request, "المحضر غير موجود")
        return redirect('elections:bureau_list')

    try:
        role_obj = request.user.role_profile
        role = role_obj.role
        if role == 'commune' and bureau.commune.id != role_obj.commune.id:
            messages.error(request, "ليس لديك صلاحية للوصول إلى هذا المحضر")
            return redirect('home')
        elif role == 'wilaya' and bureau.commune.wilaya.id != role_obj.wilaya.id:
            messages.error(request, "ليس لديك صلاحية للوصول إلى هذا المحضر")
            return redirect('home')
    except AttributeError:
        return redirect('login')

    if role == 'commune':
        communes = Commune.objects.filter(id=bureau.commune.id).select_related('wilaya')
    elif role == 'wilaya':
        communes = Commune.objects.filter(wilaya=role_obj.wilaya).select_related('wilaya').order_by('code')
    else:
        communes = Commune.objects.select_related('wilaya').order_by('wilaya__code', 'code')

    elections = Election.objects.all()

    if request.method == 'POST':
        new_commune_id = request.POST.get('commune')
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        reg_voters = request.POST.get('registered_voters', '').strip()

        if not new_commune_id or not code:
            messages.error(request, "البلدية ورمز المحضر حقول مطلوبة")
            return redirect('elections:edit_bureau', bureau_id=bureau.id)

        new_commune = Commune.objects.filter(id=new_commune_id).select_related('wilaya').first()
        if not new_commune:
            messages.error(request, "البلدية المحددة غير صحيحة")
            return redirect('elections:edit_bureau', bureau_id=bureau.id)
        
        if role == 'wilaya' and new_commune.wilaya.id != role_obj.wilaya.id:
            messages.error(request, "لا يمكنك نقل المركز إلى بلدية خارج نطاق ولايتك")
            return redirect('elections:edit_bureau', bureau_id=bureau.id)
        if role == 'commune' and new_commune.id != role_obj.commune.id:
            messages.error(request, "لا يمكنك نقل المركز إلى بلدية أخرى")
            return redirect('elections:edit_bureau', bureau_id=bureau.id)

        if BureauDeVote.objects.filter(commune=new_commune, election=bureau.election, code=code, is_deleted=False).exclude(id=bureau.id).exists():
            messages.warning(request, f"يوجد محضر آخر برمز {code} بالفعل في هذه البلدية والانتخاب")
            return redirect('elections:edit_bureau', bureau_id=bureau.id)

        reg_int = None
        if reg_voters:
            try:
                reg_int = int(reg_voters)
            except ValueError:
                messages.error(request, "عدد الناخبين المسجلين يجب أن يكون رقماً صحيحاً")
                return redirect('elections:edit_bureau', bureau_id=bureau.id)

        BureauDeVote.objects.filter(id=bureau.id).update(
            commune=new_commune, code=code, name=name if name else None, registered_voters=reg_int
        )

        messages.success(request, f" تم تحديث المحضر بنجاح (الرمز: {code})")
        return redirect('elections:bureau_list')

    return render(request, 'elections/edit_bureau.html', {
        'bureau': bureau, 'communes': communes, 'elections': elections, 'active_page': 'bureaux'
    })

@login_required
def delete_bureau(request, bureau_id):
    # 1. Fetch Bureau & Check Access
    try:
        # Removed is_deleted=False filter since we are hard deleting
        bureau = BureauDeVote.objects.select_related('commune__wilaya', 'election').get(id=bureau_id)
    except BureauDeVote.DoesNotExist:
        messages.error(request, "المحضر غير موجود أو تم حذفه مسبقاً")
        return redirect('elections:bureau_list')

    # 2. Role Scoping Check
    try:
        role_obj = request.user.role_profile
        role = role_obj.role

        if role == 'commune' and bureau.commune.id != role_obj.commune.id:
            messages.error(request, "ليس لديك صلاحية لحذف هذا المحضر")
            return redirect('home')
        elif role == 'wilaya' and bureau.commune.wilaya.id != role_obj.wilaya.id:
            messages.error(request, "ليس لديك صلاحية لحذف هذا المحضر")
            return redirect('elections:bureau_list')
    except AttributeError:
        return redirect('login')

    # 3. Handle POST (TRUE HARD DELETE)
    if request.method == 'POST':
        bureau_name = bureau.name or bureau.code
        bureau_code = bureau.code
        
        # 🧹 HARD DELETE CASCADE
        # 1. Delete StationResult first (This bypasses the RESTRICT rule)
        # Note: ListResults will be deleted automatically because they have on_delete=CASCADE
        try:
            StationResult.objects.filter(bureau=bureau).delete()
        except Exception:
            pass
            
        # 2. Hard delete the BureauDeVote itself from the database
        bureau.delete()
        
        messages.success(request, f"✅ تم حذف المحضر '{bureau_name}' (رمز: {bureau_code}) نهائياً من قاعدة البيانات.")
        return redirect('elections:bureau_list')

    # 4. Handle GET (Show confirmation page)
    return render(request, 'elections/delete_bureau_confirm.html', {
        'bureau': bureau,
        'active_page': 'bureaux'
    })

@login_required
def export_results(request):
    active_election = Election.objects.filter(status='open').order_by('-election_date').first()
    if not active_election:
        messages.error(request, "لا توجد انتخابات نشطة للتصدير")
        return redirect('home')
    
    try:
        role_obj = request.user.role_profile
        role = role_obj.role
    except:
        return redirect('login')
    
    # 🚀 OPTIMIZATION 1: Perfect Prefetching
    bureaux_qs = BureauDeVote.objects.filter(
        election=active_election, is_deleted=False
    ).select_related(
        'commune__wilaya'
    ).prefetch_related(
        'result__list_results__election_list'
    )
    
    if role == 'commune':
        bureaux_qs = bureaux_qs.filter(commune=role_obj.commune)
    elif role == 'wilaya':
        bureaux_qs = bureaux_qs.filter(commune__wilaya=role_obj.wilaya)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "النتائج"
    ws.sheet_view.rightToLeft = True
    
    headers = [
        'الولاية', 'البلدية', 'رمز المحضر', 'اسم المحضر',
        'المسجلون', 'المصوتون', 'الملغاة', 'البيضاء',
        'أصوات الحركة', 'نسبة المشاركة', 'حصة الحركة',
        'تاريخ الإدخال', 'المستخدم'
    ]
    
    header_fill = PatternFill(start_color="008A5E", end_color="008A5E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 🚀 OPTIMIZATION 2: Stream from DB & Use Prefetch Cache
    row_num = 2
    
    # .iterator() prevents Django from loading all 1541 objects into RAM at once
    for bureau in bureaux_qs.iterator(chunk_size=500):
        
        # ❌ OLD WAY (Triggers 1541 extra SQL queries): 
        # result = bureau.result.filter(is_deleted=False).first()
        
        # ✅ NEW WAY (Uses the cached prefetch data, ZERO extra queries):
        result = None
        try:
            # bureau.result.all() returns the prefetched list instantly from RAM
            for r in bureau.result.all():
                if not r.is_deleted:
                    result = r
                    break
        except Exception:
            pass
            
        if result:
            total_cast = result.total_votes_cast or 0
            registered = result.registered_voters or bureau.registered_voters or 0
            msp_votes = 0
            
            # ✅ Use prefetched list_results (ZERO extra queries)
            for lr in result.list_results.all():
                if lr.election_list.is_our_party:
                    msp_votes = lr.votes
                    break
            
            turnout = (total_cast / registered * 100) if registered > 0 else 0
            msp_share = (msp_votes / total_cast * 100) if total_cast > 0 else 0
            
            row_data = [
                bureau.commune.wilaya.name_ar,
                bureau.commune.name_ar,
                bureau.code,
                bureau.name or '',
                registered,
                total_cast,
                result.null_votes or 0,
                result.blank_votes or 0,
                msp_votes,
                f"{turnout:.1f}%",
                f"{msp_share:.1f}%",
                result.submitted_at.strftime('%Y-%m-%d %H:%M') if result.submitted_at else '',
                result.submitted_by.full_name if result.submitted_by else ''
            ]
        else:
            row_data = [
                bureau.commune.wilaya.name_ar,
                bureau.commune.name_ar,
                bureau.code,
                bureau.name or '',
                bureau.registered_voters or 0,
                0, 0, 0, 0,
                '0%', '0%',
                '', ''
            ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center' if isinstance(value, (int, float)) else 'right')
            cell.border = thin_border
            
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = f"MSP_{active_election.name_fr.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def calculate_seats_allocation(valid_votes, list_votes_dict, total_seats):
    meta = {
        'valid_votes': valid_votes, 'total_seats': total_seats, 'threshold': 0, 'quotient': 0,
        'eliminated_votes': 0, 'new_valid_votes': valid_votes,
    }
    
    if total_seats <= 0 or valid_votes <= 0 or not list_votes_dict:
        return {'meta': meta, 'lists': {}}

    threshold = valid_votes * 0.05
    meta['threshold'] = threshold

    qualified_lists = {}
    eliminated_votes = 0
    
    for lid, votes in list_votes_dict.items():
        if votes >= threshold:
            qualified_lists[lid] = votes
        else:
            eliminated_votes += votes

    meta['eliminated_votes'] = eliminated_votes
    new_valid_votes = valid_votes - eliminated_votes
    meta['new_valid_votes'] = new_valid_votes
    
    if new_valid_votes <= 0 or not qualified_lists:
        return {'meta': meta, 'lists': {}}

    quotient = math.ceil(new_valid_votes / total_seats)
    meta['quotient'] = quotient

    results = {}
    allocated_seats = 0
    remainders = []
    
    for lid, votes in qualified_lists.items():
        initial_seats = int(votes / quotient) if quotient > 0 else 0
        exact_division = votes / quotient if quotient > 0 else 0
        decimal_remainder = exact_division - initial_seats 
        
        results[lid] = {
            'seats': initial_seats, 'initial_seats': initial_seats, 'status': 'qualified',
            'remainder': decimal_remainder, 'exact_division': exact_division, 'votes': votes, 'extra_seats': 0
        }
        allocated_seats += initial_seats
        remainders.append((lid, decimal_remainder, votes))

    for lid, votes in list_votes_dict.items():
        if lid not in results:
            results[lid] = {
                'seats': 0, 'initial_seats': 0, 'status': 'eliminated', 'remainder': 0,
                'exact_division': 0, 'votes': votes, 'extra_seats': 0
            }

    remaining_seats = total_seats - allocated_seats
    
    if remaining_seats > 0 and remainders:
        remainders.sort(key=lambda x: x[1], reverse=True)
        
        seats_given_in_pass1 = min(remaining_seats, len(remainders))
        for i in range(seats_given_in_pass1):
            lid = remainders[i][0]
            results[lid]['seats'] += 1
            results[lid]['extra_seats'] += 1
            
        remaining_seats -= seats_given_in_pass1
        
        if remaining_seats > 0:
            for _ in range(remaining_seats):
                best_lid = None
                best_avg = -1
                for lid, votes in qualified_lists.items():
                    current_seats = results[lid]['seats']
                    avg = votes / (current_seats + 1)
                    if avg > best_avg:
                        best_avg = avg
                        best_lid = lid
                
                if best_lid:
                    results[best_lid]['seats'] += 1
                    results[best_lid]['extra_seats'] += 1

    meta['total_initial_seats'] = sum(r['initial_seats'] for r in results.values())
    meta['total_extra_seats'] = sum(r['extra_seats'] for r in results.values())
    meta['total_final_seats'] = sum(r['seats'] for r in results.values())

    return {'meta': meta, 'lists': results}

@login_required
def add_list_ajax(request):
    if request.method == 'POST':
        try:
            role_obj = request.user.role_profile
            role = role_obj.role
        except AttributeError:
            return JsonResponse({'success': False, 'error': 'غير مصرح'})

        election_id = request.POST.get('election_id')
        bureau_id = request.POST.get('bureau_id')
        name = request.POST.get('name', '').strip()
        
        # 🛡️ SECURITY ENFORCEMENT: 
        # Only National/Super Admins can create National lists or mark them as MSP
        if role in ['national', 'super_admin']:
            is_our_party = request.POST.get('is_our_party') == 'true'
            is_national = request.POST.get('is_national') == 'true'
        else:
            # Wilaya and Commune reps can ONLY create local, standard lists
            is_our_party = False
            is_national = False
            
        if not name or not election_id or not bureau_id:
            return JsonResponse({'success': False, 'error': 'بيانات ناقصة'})
            
        election = Election.objects.filter(id=election_id).first()
        bureau = BureauDeVote.objects.select_related('commune__wilaya').filter(id=bureau_id, is_deleted=False).first()
        
        if not election or not bureau:
            return JsonResponse({'success': False, 'error': 'الانتخاب أو المحضر غير موجود'})
            
        # If it's a local list (created by wilaya/commune), tie it to their specific wilaya
        wilaya = None if is_national else bureau.commune.wilaya
        
        if ElectionList.objects.filter(election=election, name_ar=name, wilaya=wilaya).exists():
            return JsonResponse({'success': False, 'error': 'هذه القائمة موجودة مسبقاً'})
            
        max_order = ElectionList.objects.filter(election=election).aggregate(Max('display_order'))['display_order__max'] or 0
        
        new_list = ElectionList.objects.create(
            election=election, 
            name_ar=name, 
            is_our_party=is_our_party, 
            wilaya=wilaya,
            display_order=max_order + 1
        )
        
        return JsonResponse({
            'success': True, 
            'id': new_list.id, 
            'name': new_list.name_ar,
            'is_national': is_national
        })
        
    return JsonResponse({'success': False})

@login_required
def update_wilaya_seats(request, wilaya_id):
    if request.method == 'POST':
        try:
            role_obj = request.user.role_profile
            role = role_obj.role
        except AttributeError:
            return JsonResponse({'success': False, 'error': 'غير مصرح'})

        wilaya = get_object_or_404(Wilaya, id=wilaya_id)

        if role not in ['national', 'super_admin']:
            return JsonResponse({'success': False, 'error': '⛔ ليس لديك صلاحية تعديل المقاعد. هذه الصلاحية للمسؤول الوطني فقط.'})
            
        seats = request.POST.get('seats')
        if not seats or int(seats) < 1:
            return JsonResponse({'success': False, 'error': 'عدد غير صحيح'})

        wilaya.total_seats = int(seats)
        wilaya.save(update_fields=['total_seats'])
        
        # 🧹 PILLAR 2: INSTANTLY CLEAR CACHE WHEN SEATS CHANGE
        active_election = Election.objects.filter(status='open').order_by('-election_date').first()
        if active_election:
            cache.delete(f"wilaya_math_{wilaya.id}_{active_election.id}")
            cache.delete(f"wilaya_detail_math_{wilaya.id}_{active_election.id}")
            cache.delete(f"national_math_{active_election.id}")
            
        return JsonResponse({'success': True, 'seats': wilaya.total_seats})
        
    return JsonResponse({'success': False})

@login_required
def audit_log(request):
    try:
        role_obj = request.user.role_profile
        role = role_obj.role
    except AttributeError:
        messages.error(request, "غير مصرح لك بالوصول")
        return redirect('home')
    
    if role not in ['national', 'super_admin']:
        messages.error(request, "⛔ غير مصرح لك بالوصول إلى سجل التدقيق")
        return redirect('home')
    
    user_filter = request.GET.get('user')
    action_filter = request.GET.get('action')
    table_filter = request.GET.get('table')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search', '').strip()
    
    logs = ResultAuditLog.objects.select_related('changed_by').all().order_by('-timestamp')
    
    if user_filter:
        logs = logs.filter(changed_by_id=user_filter)
    if action_filter:
        logs = logs.filter(action=action_filter)
    if table_filter:
        logs = logs.filter(table_name=table_filter)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    if search_query:
        logs = logs.filter(
            Q(notes__icontains=search_query) | Q(record_id__icontains=search_query) |
            Q(changed_by__full_name__icontains=search_query) | Q(changed_by__username__icontains=search_query)
        )
    
    total_logs = ResultAuditLog.objects.count()
    today_logs = ResultAuditLog.objects.filter(timestamp__date=timezone.now().date()).count()
    critical_actions = ResultAuditLog.objects.filter(action='delete').count()
    
    CustomUser = get_user_model()
    all_users = CustomUser.objects.filter(role_profile__isnull=False).select_related('role_profile').order_by('full_name')
    action_choices = ResultAuditLog.ACTION_CHOICES
    table_choices = ResultAuditLog.objects.values_list('table_name', flat=True).distinct().order_by('table_name')
    
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj, 'total_logs': total_logs, 'today_logs': today_logs, 'critical_actions': critical_actions,
        'all_users': all_users, 'action_choices': action_choices, 'table_choices': table_choices,
        'user_filter': user_filter, 'action_filter': action_filter, 'table_filter': table_filter,
        'date_from': date_from, 'date_to': date_to, 'search_query': search_query, 'active_page': 'audit_log',
    }
    
    return render(request, 'elections/audit_log.html', context)

@login_required
def get_audit_details(request, log_id):
    try:
        role = request.user.role_profile.role
        if role not in ['national', 'super_admin']:
            return JsonResponse({'error': 'غير مصرح'}, status=403)
    except:
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    log = get_object_or_404(ResultAuditLog, id=log_id)
    time_str = log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else ''
    
    return JsonResponse({
        'table': log.table_name, 'record_id': log.record_id, 'action': log.get_action_display(),
        'user': log.changed_by.full_name if log.changed_by else 'النظام', 'time': time_str,
        'notes': log.notes, 'old_values': log.old_values, 'new_values': log.new_values
    })